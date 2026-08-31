#!/usr/bin/env python3
"""Document-style Excel helpers for Huawei D-MIMO deployment packs."""

from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage


# Huawei-inspired document palette
RED = "C7000B"
DARK = "1C1C1C"
NAVY = "243447"
SLATE = "4A5568"
GOLD = "B8860B"
WHITE = "FFFFFF"
IVORY = "FBF7F2"
LIGHT = "F4F1EC"
ALT = "F8EEEC"
BLUE_BG = "E8F1F8"
GREEN_BG = "E8F5E9"
YELLOW_BG = "FFF8E1"
ORANGE_BG = "FFF3E0"
RED_BG = "FDECEC"
STEP_BG = "EEF4FA"
MML_BG = "1E2430"
MML_FG = "E8EEF4"
TBL_HEAD = "C7000B"
TBL_HEAD2 = "243447"

THIN = Border(
    left=Side(style="thin", color="C9C2B8"),
    right=Side(style="thin", color="C9C2B8"),
    top=Side(style="thin", color="C9C2B8"),
    bottom=Side(style="thin", color="C9C2B8"),
)
MED_RED = Border(
    left=Side(style="medium", color=RED),
    right=Side(style="medium", color=RED),
    top=Side(style="medium", color=RED),
    bottom=Side(style="medium", color=RED),
)

NCOLS = 12  # A-L
COL_WIDTHS = {
    "A": 4.2,
    "B": 22.0,
    "C": 18.0,
    "D": 22.0,
    "E": 16.0,
    "F": 16.0,
    "G": 16.0,
    "H": 14.0,
    "I": 14.0,
    "J": 14.0,
    "K": 14.0,
    "L": 18.0,
}


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(name="Calibri", size=11, bold=False, italic=False, color=DARK, underline=None):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color, underline=underline)


def align(h="left", v="center", wrap=True, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def row_height_for(text: str, chars_per_line: int = 110, min_h: float = 18, max_h: float = 220) -> float:
    if not text:
        return min_h
    lines = 0
    for para in str(text).split("\n"):
        para = para or " "
        lines += max(1, (len(para) + chars_per_line - 1) // chars_per_line)
    return min(max_h, max(min_h, lines * 15.0 + 8))


class DocBook:
    def __init__(self, title: str, subtitle: str, doc_code: str):
        self.wb = Workbook()
        self.title = title
        self.subtitle = subtitle
        self.doc_code = doc_code
        self._first = True
        self.toc = []  # (sheet_name, heading)

    def sheet(self, name: str, heading: str, subheading: str = "", chapter: str = "") -> "DocSheet":
        if self._first:
            ws = self.wb.active
            ws.title = name[:31]
            self._first = False
        else:
            ws = self.wb.create_sheet(name[:31])
        ds = DocSheet(ws, self, heading, subheading, chapter)
        self.toc.append((ws.title, heading, subheading or chapter))
        return ds

    def save(self, path: str):
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        self.wb.save(path)


class DocSheet:
    def __init__(self, ws, book: DocBook, heading: str, subheading: str, chapter: str):
        self.ws = ws
        self.book = book
        self.r = 1
        self.heading = heading
        for col, w in COL_WIDTHS.items():
            ws.column_dimensions[col].width = w
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.5, header=0.25, footer=0.25)
        ws.print_options.horizontalCentered = True
        ws.oddHeader.left.text = f"&8{book.doc_code}"
        ws.oddHeader.right.text = "&8eRAN22.1 01 | 2026-03-10"
        ws.oddFooter.left.text = "&8Huawei D-MIMO (TDD) Feature Pack  |  Internal deployment use"
        ws.oddFooter.right.text = "&8Page &P of &N"
        ws.print_title_rows = "1:3"
        ws.freeze_panes = "A4"
        self._banner(heading, subheading, chapter)

    def _merge(self, r1, c1, r2, c2):
        if c2 > c1 or r2 > r1:
            self.ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def _cell(self, r, c, value=None, fnt=None, fil=None, al=None, border=None):
        cell = self.ws.cell(r, c)
        if isinstance(cell, MergedCell):
            return cell
        if value is not None:
            cell.value = value
        if fnt:
            cell.font = fnt
        if fil:
            cell.fill = fil
        if al:
            cell.alignment = al
        if border:
            cell.border = border
        return cell

    def _band(self, r, c1, c2, value, fnt, fil, al, border=None):
        """Style a row span, write value on origin, then merge."""
        for c in range(c1, c2 + 1):
            self._cell(r, c, None, fnt=fnt, fil=fil, al=al, border=border)
        self._cell(r, c1, value, fnt=fnt, fil=fil, al=al, border=border)
        self._merge(r, c1, r, c2)

    def _banner(self, heading, subheading, chapter):
        ws = self.ws
        for c in range(1, NCOLS + 1):
            self._cell(1, c, "", fil=fill(RED))
        self._merge(1, 1, 1, NCOLS)
        self._cell(
            1, 1,
            "HUAWEI  eRAN  |  D-MIMO (TDD) Feature Parameter Description  |  Issue 01  |  2026-03-10",
            fnt=font(size=9, bold=True, color=WHITE),
            fil=fill(RED),
            al=align("left", "center"),
        )
        ws.row_dimensions[1].height = 16

        for c in range(1, NCOLS + 1):
            self._cell(2, c, "", fil=fill(NAVY))
        self._merge(2, 1, 2, NCOLS)
        title = heading if not chapter else f"{chapter}    {heading}"
        self._cell(2, 1, title, fnt=font(size=16, bold=True, color=WHITE), fil=fill(NAVY), al=align("left", "center"))
        ws.row_dimensions[2].height = 28

        for c in range(1, NCOLS + 1):
            self._cell(3, c, "", fil=fill(IVORY))
        self._merge(3, 1, 3, NCOLS)
        self._cell(
            3, 1,
            subheading or self.book.subtitle,
            fnt=font(size=10, italic=True, color=SLATE),
            fil=fill(IVORY),
            al=align("left", "center"),
        )
        ws.row_dimensions[3].height = 18
        self.r = 4

    def spacer(self, h=8):
        self.ws.row_dimensions[self.r].height = h
        self.r += 1

    def h2(self, text: str):
        self.spacer(6)
        self._band(self.r, 1, NCOLS, text, font(size=13, bold=True, color=WHITE), fill(RED), align("left", "center"))
        self.ws.row_dimensions[self.r].height = 22
        self.r += 1

    def h3(self, text: str):
        self.spacer(4)
        self._band(self.r, 1, NCOLS, text, font(size=11, bold=True, color=WHITE), fill(NAVY), align("left", "center"))
        self.ws.row_dimensions[self.r].height = 20
        self.r += 1

    def h4(self, text: str):
        self._cell(self.r, 1, "", fil=fill(GOLD))
        self._band(self.r, 2, NCOLS, text, font(size=11, bold=True, color=RED), None, align("left", "center"))
        self.ws.row_dimensions[self.r].height = 18
        self.r += 1

    def para(self, text: str, bg=None, italic=False, bold=False, color=DARK):
        fil = fill(bg) if bg else None
        self._band(
            self.r, 2, NCOLS, text,
            font(size=11, bold=bold, italic=italic, color=color),
            fil,
            align("left", "top"),
        )
        self.ws.row_dimensions[self.r].height = row_height_for(text)
        self.r += 1

    def note(self, text: str, kind="NOTE"):
        colors = {
            "NOTE": (YELLOW_BG, GOLD, "NOTE"),
            "NOTICE": (ORANGE_BG, "D35400", "NOTICE"),
            "WARNING": (RED_BG, RED, "WARNING"),
            "TIP": (GREEN_BG, "1B7A3D", "TIP"),
            "REVIEW": (BLUE_BG, NAVY, "REVIEW"),
            "TRIAL": (RED_BG, RED, "TRIAL FEATURE"),
        }
        bg, fg, label = colors.get(kind, (YELLOW_BG, GOLD, kind))
        self._band(
            self.r, 2, NCOLS, f"{label}:  {text}",
            font(size=10, italic=True, color=fg),
            fill(bg),
            align("left", "top"),
            border=THIN,
        )
        self.ws.row_dimensions[self.r].height = row_height_for(f"{label}:  {text}", 100)
        self.r += 1

    def bullets(self, items, numbered=False):
        for i, item in enumerate(items, 1):
            prefix = f"{i}.  " if numbered else "•  "
            self._band(self.r, 2, NCOLS, prefix + str(item), font(size=11), None, align("left", "top"))
            self.ws.row_dimensions[self.r].height = row_height_for(str(item), 105)
            self.r += 1

    def step(self, num, title, body="", action=None):
        for c in range(1, NCOLS + 1):
            self._cell(self.r, c, None, fil=fill(STEP_BG), border=THIN)
        self._cell(
            self.r, 1, f"STEP {num}",
            fnt=font(size=10, bold=True, color=WHITE),
            fil=fill(RED),
            al=align("center", "center"),
            border=THIN,
        )
        self._cell(
            self.r, 2, title,
            fnt=font(size=11, bold=True, color=NAVY),
            fil=fill(STEP_BG),
            al=align("left", "center"),
            border=THIN,
        )
        self._merge(self.r, 2, self.r, NCOLS)
        self.ws.row_dimensions[self.r].height = 22
        self.r += 1
        if body:
            self.para(body)
        if action:
            self.para("Action:  " + action, bg=GREEN_BG, bold=True)

    def kv_table(self, rows, headers=None):
        """Simple 2-col key/value starting at column B."""
        data = []
        if headers:
            data.append(headers)
        data.extend(rows)
        self.table(data, col_start=2, col_span=None)

    def table(self, rows, col_start=2, widths_hint=None, header_fill=TBL_HEAD):
        if not rows:
            return
        ncols = max(len(r) for r in rows)
        col_end = min(NCOLS, col_start + ncols - 1)
        # if last columns remain, merge last data col into remaining
        extra = NCOLS - col_end
        for i, row in enumerate(rows):
            is_head = i == 0
            bg = header_fill if is_head else (ALT if i % 2 == 0 else WHITE)
            fg = WHITE if is_head else DARK
            f = font(size=9 if not is_head else 10, bold=True if is_head else False, color=fg)
            cells = list(row) + [""] * (ncols - len(row))
            for j in range(ncols):
                c = col_start + j
                if c > NCOLS:
                    break
                # merge last column across leftover columns on first row only handled per cell
                val = cells[j]
                if j == ncols - 1 and col_start + j < NCOLS:
                    for cc in range(c, NCOLS + 1):
                        self._cell(self.r, cc, None, fnt=f, fil=fill(bg), al=align("left", "top"), border=THIN)
                    self._cell(self.r, c, val, fnt=f, fil=fill(bg), al=align("left", "top"), border=THIN)
                    self._merge(self.r, c, self.r, NCOLS)
                else:
                    self._cell(self.r, c, val, fnt=f, fil=fill(bg), al=align("left", "top"), border=THIN)
            text_join = " ".join(str(x) for x in cells)
            self.ws.row_dimensions[self.r].height = row_height_for(text_join, chars_per_line=28 * ncols, min_h=20, max_h=90)
            self.r += 1
        self.spacer(4)

    def mml(self, title, commands: str):
        self.h4(title)
        # split into lines, each as a code row
        for line in commands.strip("\n").split("\n"):
            color = "8FBF8F" if line.strip().startswith("//") else MML_FG
            self._band(
                self.r, 2, NCOLS, line if line.strip() else " ",
                font(name="Consolas", size=9, color=color),
                fill(MML_BG),
                align("left", "center"),
            )
            self.ws.row_dimensions[self.r].height = 16 if len(line) < 120 else 28
            self.r += 1
        self.spacer(4)

    def figure(self, path: str, caption: str, max_width_px: int = 720):
        if not os.path.exists(path):
            self.note(f"Figure missing: {path}", "WARNING")
            return
        self.h4(caption)
        im = PILImage.open(path)
        w, h = im.size
        scale = min(1.0, max_width_px / float(w))
        disp_w = int(w * scale)
        disp_h = int(h * scale)
        img = XLImage(path)
        img.width = disp_w
        img.height = disp_h
        # Place at column B
        self.ws.add_image(img, f"B{self.r}")
        # Excel row height is in points (~0.75 px)
        rows_needed = max(6, int((disp_h / 13.0) + 1))
        for i in range(rows_needed):
            self.ws.row_dimensions[self.r].height = 13
            self.r += 1
        self._band(
            self.r, 2, NCOLS, caption,
            font(size=9, italic=True, color=SLATE),
            None,
            align("center", "center"),
        )
        self.ws.row_dimensions[self.r].height = 16
        self.r += 1
        self.spacer(6)

    def check_row(self, no, item, criteria, owner="RAN Eng.", status="☐"):
        return [str(no), item, criteria, owner, status]

    def footer_end(self):
        self.spacer(10)
        self._band(
            self.r, 2, NCOLS,
            "Source: Huawei eRAN D-MIMO (TDD) Feature Parameter Description, Issue 01 (2026-03-10). "
            "This workbook is a deployment-oriented restatement of that document. Feature gains depend on the live scenario; contact Huawei professional service for cluster planning and optimization.",
            font(size=8, italic=True, color=SLATE),
            None,
            align("left", "top"),
        )
        self.ws.row_dimensions[self.r].height = 36
        self.r += 1
        self.ws.auto_filter.ref = None
        # print area
        self.ws.print_area = f"A1:L{self.r}"
        return self.r
